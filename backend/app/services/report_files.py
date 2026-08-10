import csv

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BRAND_DARK = colors.HexColor("#051F20")
BRAND_ACCENT = colors.HexColor("#235347")
BRAND_SOFT = colors.HexColor("#DAF1DE")


def generate_pdf(data: dict, store_name: str, out_path: str):
    doc = SimpleDocTemplate(out_path, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("RITitle", parent=styles["Title"], textColor=BRAND_DARK, fontSize=20)
    meta_style = ParagraphStyle("RIMeta", parent=styles["Normal"], textColor=colors.grey, fontSize=10)
    section_style = ParagraphStyle("RISection", parent=styles["Heading2"], textColor=BRAND_DARK, spaceBefore=14)

    story = [
        Paragraph(store_name, meta_style),
        Paragraph(data["title"], title_style),
        Paragraph(f"Period: {data['period']}", meta_style),
        Spacer(1, 16),
    ]

    # --- Summary KPI table ---
    summary_rows = [[k, _fmt(v)] for k, v in data["summary"].items()]
    summary_table = Table(summary_rows, colWidths=[2.8 * inch, 2.8 * inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_SOFT),
        ("TEXTCOLOR", (0, 0), (0, -1), BRAND_DARK),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#c7ddd5")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(summary_table)

    # --- Detail table ---
    if data["table"]:
        story.append(Paragraph(data["table_title"], section_style))
        headers = list(data["table"][0].keys())
        rows = [headers] + [[_fmt(row[h]) for h in headers] for row in data["table"][:200]]
        col_width = 6.8 * inch / len(headers)
        detail_table = Table(rows, colWidths=[col_width] * len(headers), repeatRows=1)
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f9f6")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dfeee8")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(detail_table)
        if len(data["table"]) > 200:
            story.append(Spacer(1, 8))
            story.append(Paragraph(f"...and {len(data['table']) - 200} more rows (see the Excel/CSV export for the full data).", meta_style))
    else:
        story.append(Spacer(1, 16))
        story.append(Paragraph("No data available for this period.", meta_style))

    doc.build(story)
    return out_path


def generate_excel(data: dict, store_name: str, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill(start_color="051F20", end_color="051F20", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="051F20")

    ws["A1"] = store_name
    ws["A2"] = data["title"]
    ws["A2"].font = title_font
    ws["A3"] = f"Period: {data['period']}"

    row = 5
    ws.cell(row=row, column=1, value="Metric").font = header_font
    ws.cell(row=row, column=2, value="Value").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2).fill = header_fill
    for k, v in data["summary"].items():
        row += 1
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=v)

    if data["table"]:
        ws2 = wb.create_sheet(data["table_title"][:31])
        headers = list(data["table"][0].keys())
        for col, h in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r, item in enumerate(data["table"], start=2):
            for col, h in enumerate(headers, start=1):
                ws2.cell(row=r, column=col, value=item[h])
        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[chr(64 + col)].width = 18

    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 22

    wb.save(out_path)
    return out_path


def generate_csv(data: dict, out_path: str):
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([data["title"]])
        writer.writerow([f"Period: {data['period']}"])
        writer.writerow([])
        writer.writerow(["Metric", "Value"])
        for k, v in data["summary"].items():
            writer.writerow([k, v])
        writer.writerow([])
        if data["table"]:
            headers = list(data["table"][0].keys())
            writer.writerow(headers)
            for row in data["table"]:
                writer.writerow([row[h] for h in headers])
    return out_path


def _fmt(value):
    if isinstance(value, float):
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)
